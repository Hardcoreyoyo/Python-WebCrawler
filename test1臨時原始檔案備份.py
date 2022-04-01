import glob
import os
import sys
import time
from os import listdir
from os.path import isfile, isdir, join

import urllib
from urllib.request import urlretrieve

import openpyxl
import requests
from bs4 import BeautifulSoup
from bs4 import BeautifulSoup
import pandas as pd

from openpyxl.utils.cell import get_column_letter

import xlsxwriter
import openpyxl
from openpyxl import load_workbook, drawing
from openpyxl.drawing.image import Image
from openpyxl.drawing import image
from openpyxl.workbook import Workbook
# 輸出或寫入excel都要import這模組

from pandas import DataFrame

headers = {
    'user-agent': 'Mozilla/5.0 (Macintosh Intel Mac OS X 10_13_4) AppleWebKit/537.36 (KHTML, like Gecko) '
                  'Chrome/66.0.3359.181 Safari/537.36'}

res = requests.get('https://rate.bot.com.tw/xrt?Lang=zh-TW', headers=headers)
WebAllData = res.text
soup = BeautifulSoup(WebAllData, 'html.parser')

Currency = []  # 取得幣種類名稱
Cash_Exchange_Rate_Buy = []  # 取得現金匯率_本行現金買入
Cash_Exchange_Rate_Sell = []  # 取得現金匯率_本行現金賣出
Cash_Exchange_Rate_BuyNow = []  # 取得即期匯率_本行即期買入
Cash_Exchange_Rate_SellNow = []  # 取得即期匯率_本行即期賣出
DataInTable = {"現金買入": Cash_Exchange_Rate_Buy,
               "現金賣出": Cash_Exchange_Rate_Sell,
               "即期買入": Cash_Exchange_Rate_BuyNow,
               "即期賣出": Cash_Exchange_Rate_SellNow}

for link in soup.find_all(class_='hidden-phone print_show'):
    organize = link.text.strip()
    Currency.append(organize)

for buy in soup.find_all(attrs={"data-table": "本行現金買入",
                                "class": "rate-content-cash text-right print_hide"}):
    Cash_Exchange_Rate_Buy.append(buy.text.strip())

for sell in soup.find_all(attrs={"data-table": "本行現金賣出",
                                 "class": "rate-content-cash text-right print_hide"}):
    Cash_Exchange_Rate_Sell.append(sell.text.strip())

for buy_now in soup.find_all(attrs={"data-table": "本行即期買入",
                                    "class": "text-right display_none_print_show print_width"}):
    Cash_Exchange_Rate_BuyNow.append(buy_now.text.strip())

for sell_now in soup.find_all(attrs={"data-table": "本行即期賣出",
                                     "class": "text-right display_none_print_show print_width"}):
    Cash_Exchange_Rate_SellNow.append(sell_now.text.strip())

# images = []
# for i in soup.find_all(attrs={"title": "幣別國旗"}):
#     if "https://rate.bot.com.tw" not in i.get('src'):  # 判断URL是否完整
#         images.append('https://rate.bot.com.tw' + i.get('src'))


# images = []
# for i in soup.find_all('img', title='幣別國旗'):
#     c = i.get('src')
#     if "https://rate.bot.com.tw" not in c:  # 判断URL是否完整
#         images.append('https://rate.bot.com.tw' + c)


# for i in soup.tbody.children:
#     print(i)

# i = 0
# k = 0
# for j in soup.find_all('img', title='幣別國旗'):  # 存圖片到電腦資料夾
#     i += 1
#     k += 1
#     c = j.get('src')
#
#     if "https://rate.bot.com.tw" not in c:  # 判断URL是否完整
#         c = 'https://rate.bot.com.tw' + c
#
#     filename = 'C:\\Users\\Hardcoreyoyo\\Desktop\\Python\\WebCrawlerLearn\\WebCrawlerTest1\\images\\' \
#                + 'photoS' + str(i) + '.png'
#
#     with open(filename, 'w'):
#         urllib.request.urlretrieve(c, filename)
#
#     print('存了' + str(k) + '張')
#
# print('處理結束')


# i = 0  # 用插入的方式excel
# k = 0
# g = 0
# p = 0
# InData = pd.ExcelWriter('testtest1111.xlsx', engine='xlsxwriter')
# workbook = InData.book
# worksheet = workbook.add_worksheet()
# for j in soup.find_all('img', title='幣別國旗'):  # 存圖片到電腦資料夾
#     i += 1
#     p += 5
#     k += 1
#     g += 1
#     c = j.get('src')
#
#     if "https://rate.bot.com.tw" not in c:  # 判断URL是否完整
#         c = 'https://rate.bot.com.tw' + c
#
#     filename = 'C:\\Users\\Hardcoreyoyo\\Desktop\\Python\\WebCrawlerLearn\\WebCrawlerTest1\\images\\' \
#                + 'photoS' + str(i) + '.png'
#
#     with open(filename, 'w'):
#         urllib.request.urlretrieve(c, filename)
#
#     worksheet.insert_image('A' + str(p),
#                            'C:\\Users\\Hardcoreyoyo\\Desktop\\Python\\WebCrawlerLearn\\WebCrawlerTest1\\images\\' + 'photoS' + str(
#                                g) + '.png')
#
#     print('存了' + str(k) + '張')
#
# InData.save()
# print('處理結束')


df = pd.DataFrame(DataInTable, index=Currency)
InData = pd.ExcelWriter('C:\\Users\\Hardcoreyoyo\\Desktop\\Python\\WebCrawlerLearn\\WebCrawlerTest1\\testtest7777', engine='xlsxwriter')
df.to_excel(InData, sheet_name='sheet1')
# workbook = InData.book
# worksheet = InData.sheets['Sheet1']
# i=0
# for f
# worksheet.insert_image('D3',
#                        'C:\\Users\\Hardcoreyoyo\\Desktop\\Python\\WebCrawlerLearn\\WebCrawlerTest1\\images\\' + 'photoS' + str(i) + '.png')
# InData.save()

# with pd.ExcelWriter('test.xlsx') as iNData: #寫入excel
#     df.to_excel(iNData, sheet_name="stocks")


# 插入圖片第二種方法
# wb = load_workbook('C:\\Users\\Hardcoreyoyo\\Desktop\\Python\\WebCrawlerLearn\\WebCrawlerTest1\\text.xlsx')  # 把檔案先讀出來
# print('讀取成功')
# print(wb)
# ws = wb.active  # 要把圖檔加進第一個sheet
# print('加載SHEET成功')
# imagesPath = 'C:\\Users\\Hardcoreyoyo\\Desktop\\Python\\WebCrawlerLearn\\WebCrawlerTest1\\images\\photo1.png'
# imagePath = os.path.join(imagesPath)
# img = Image(imagePath)
# img.width, img.height = (300, 300)
# ws.add_image(img, 'A1')
# wb.save('C:\\Users\\Hardcoreyoyo\\Desktop\\Python\\WebCrawlerLearn\\WebCrawlerTest1\\text.xlsx')  # 不要忘記save


# 插入圖片第三種方法 並讀取時按照建立日期
# 要記得import 這模組 from openpyxl.drawing.image import Image

# col = 0
# wb = load_workbook('C:\\Users\\Hardcoreyoyo\\Desktop\\Python\\WebCrawlerLearn\\WebCrawlerTest1\\text.xlsx')  # 把檔案先讀出來
# print('讀取成功')
# ws = wb.worksheets[0]  # 要把圖檔加進第一個sheet
# print('加載SHEET成功')
# # 使用glob套件做讀檔得動作，從一個資料夾裡把每一個檔案讀出來。
# # 使用os套件，在讀檔的時候，從時間最早的檔案先讀，避免順序不對
# searchedfiles = sorted(
#     glob.glob("C:\\Users\\Hardcoreyoyo\\Desktop\\Python\\WebCrawlerLearn\\WebCrawlerTest1\\images/*.png"),
#     key=os.path.getmtime)
# for fn in searchedfiles:
#     img = openpyxl.drawing.image.Image(fn)  # create image instances
#     c = str(col + 2)
#     ws.add_image(img, 'A' + c)
#     col = col + 1
# wb.save('C:\\Users\\Hardcoreyoyo\\Desktop\\Python\\WebCrawlerLearn\\WebCrawlerTest1\\text.xlsx')  # 不要忘記save


# col_num = 1
# col = 1
# wb = load_workbook('C:\\Users\\Hardcoreyoyo\\Desktop\\Python\\WebCrawlerLearn\\WebCrawlerTest1\\text.xlsx')  # 把檔案先讀出來
# print('讀取成功')
# ws = wb.worksheets[0]  # 要把圖檔加進第一個sheet
# print('加載SHEET成功')


# searchedfiles = sorted(
#     glob.glob("C:\\Users\\Hardcoreyoyo\\Desktop\\Python\\WebCrawlerLearn\\WebCrawlerTest1\\images/*.png"),
#     key=os.path.getmtime)
#
# Count_List = len(searchedfiles)
# for n in range(0, Count_List):
#     for fn in searchedfiles:
#         img = openpyxl.drawing.image.Image(fn)  # create image instances
#         Get_Col_Num = get_column_letter(col_num)
#         Sum_Col_Num = str(Get_Col_Num) + str(col)
#         ws.add_image(img, Sum_Col_Num)
#
#         print(Sum_Col_Num)
#         col_num += 1
#
# wb.save('C:\\Users\\Hardcoreyoyo\\Desktop\\Python\\WebCrawlerLearn\\WebCrawlerTest1\\text.xlsx')
# wb.close()








# # 利用時間函數創立資料夾, 並檢查資料夾是否存在
# localtime = time.localtime(time.time())
# Time_For_Dir = str(localtime[0]) + str(localtime[1]) + str(localtime[2]) + str(localtime[3]) + str(localtime[4]) + str(
#     localtime[5])
#
# mypath = "C:\\Users\\Hardcoreyoyo\\Desktop\\Python\\WebCrawlerLearn\\WebCrawlerTest1\\images"
# path = "C:\\Users\\Hardcoreyoyo\\Desktop\\Python\\WebCrawlerLearn\\WebCrawlerTest1\\images\\" + Time_For_Dir
# files = listdir(mypath)
#
# for f in files:
#     if f == Time_For_Dir:
#         print('發現相同資料夾中止')
#         sys.exit(1)
# try:
#     os.mkdir(path)
# except FileExistsError:
#     print('發現相同資料夾中止')
# print('沒有發現相同資料夾')
# print('創立資料夾:' + Time_For_Dir)
















# 按照建立順序讀取資料夾,按照建立順序讀取圖片插入excel
# wb = load_workbook('C:\\Users\\Hardcoreyoyo\\Desktop\\Python\\WebCrawlerLearn\\WebCrawlerTest1\\text.xlsx')  # 把檔案先讀出來
# print('讀取成功')
# ws = wb.worksheets[0]  # 要把圖檔加進第一個sheet
# print('加載SHEET成功')
#
# Dir_Sort_Path = glob.glob("C:\\Users\\Hardcoreyoyo\\Desktop\\Python\\WebCrawlerLearn\\WebCrawlerTest1\\images/*")
# Dir_Sort_Path_Sorted = sorted(Dir_Sort_Path, key=os.path.getctime)
# col = 1
#
# for Dir_Select in Dir_Sort_Path_Sorted:
#
#     searchedfiles = sorted(
#         glob.glob(Dir_Select + "\\/*.png"),
#         key=os.path.getctime)
#     col_num = 1
#
#     for fn in searchedfiles:
#         img = openpyxl.drawing.image.Image(fn)  # create image instances
#         Get_Col_Num = get_column_letter(col_num)
#         Sum_Col_Num = str(Get_Col_Num) + str(col)
#         ws.add_image(img, Sum_Col_Num)
#
#         print(Sum_Col_Num)
#         col_num += 2
#
#     col += 1
#
# wb.save('C:\\Users\\Hardcoreyoyo\\Desktop\\Python\\WebCrawlerLearn\\WebCrawlerTest1\\text.xlsx')
# wb.close()
# print('執行成功')
